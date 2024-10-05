import openpyxl
from openpyxl import Workbook

def add_data_to_excel(file_path, data, sheet_name="Sheet1"):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()
    
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    
    for row in data:
        sheet.append(row)
    
    workbook.save(file_path)
    print(f"Data added to {file_path} successfully.")

if __name__ == "__main__":
    file_path = "data.xlsx"
    data = [
        ["Name", "Email", "Age"],
        ["Alice", "alice@example.com", 30],
        ["Bob", "bob@example.com", 25],
        ["Charlie", "charlie@example.com", 35],
    ]
    add_data_to_excel(file_path, data)
  
